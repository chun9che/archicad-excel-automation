from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from archicad_excel.mapping import map_records
from archicad_excel.parser import parse_archicad_txt
from archicad_excel.qc import run_qc
from archicad_excel.template import load_yaml
from archicad_excel.writer import write_workbook


class EngineTests(unittest.TestCase):
    def test_parser_maps_kellerabteil_by_stable_ka_identifier(self):
        cfg = load_yaml(ROOT / "configs" / "example_project.yaml")
        records = parse_archicad_txt(ROOT / "examples" / "sanitized" / "archicad_export.txt", cfg)
        keller = [record for record in records if record.room_type == "Kellerabteil"]

        self.assertEqual([record.room_id for record in keller], ["KA1", "KA2", "KA3"])
        self.assertEqual([record.unit for record in keller], ["TOP 01", "TOP 02", "TOP 03"])

    def test_mapping_and_writer_preserve_template_formula(self):
        cfg = load_yaml(ROOT / "configs" / "example_project.yaml")
        template_cfg = load_yaml(ROOT / "configs" / "example_template.yaml")
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            template = tmp_path / "template.xlsx"
            output = tmp_path / "output.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Topographie"
            ws["D5"] = "Top"
            ws["E5"] = "Wohnfläche"
            ws["F5"] = "Balkon"
            ws["G5"] = "Loggia"
            ws["H5"] = "Kellerabteil / Einlagerungsraum"
            for idx, unit in enumerate(["TOP 01", "TOP 02", "TOP 03"], start=6):
                ws[f"D{idx}"] = unit
            ws["H9"] = "=SUM(H6:H8)"
            wb.save(template)

            records = parse_archicad_txt(ROOT / "examples" / "sanitized" / "archicad_export.txt", cfg)
            mapped, unmapped = map_records(records, cfg)
            result = write_workbook(template, output, mapped, template_cfg)
            qc = run_qc(records, mapped, unmapped)

            out = openpyxl.load_workbook(output, data_only=False)
            ws_out = out["Topographie"]
            self.assertEqual(result["missing_targets"], [])
            self.assertEqual(qc["status"], "PASS")
            self.assertEqual(ws_out["H6"].value, 3.1)
            self.assertEqual(ws_out["H7"].value, 2.95)
            self.assertEqual(ws_out["H8"].value, 4.05)
            self.assertEqual(ws_out["H9"].value, "=SUM(H6:H8)")


if __name__ == "__main__":
    unittest.main()
