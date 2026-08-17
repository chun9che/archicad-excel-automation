from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from shutil import copyfile
from typing import Any

import openpyxl

from .models import MappedRecord


def write_workbook(
    template_path: str | Path,
    output_path: str | Path,
    mapped_records: list[MappedRecord],
    template_profile: dict[str, Any],
) -> dict[str, Any]:
    """Copy a template workbook and write mapped values into configured cells."""

    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copyfile(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    worksheet_name = template_profile["template"]["worksheet"]
    ws = wb[worksheet_name]
    table = template_profile["template"]["unit_table"]
    fields = template_profile["template"]["fields"]
    first_row = int(table["first_data_row"])
    last_row = int(table.get("last_data_row") or template_profile["template"].get("total_row", ws.max_row) - 1)
    unit_column = table["unit_column"]

    row_by_unit = {
        str(ws[f"{unit_column}{row}"].value).strip(): row
        for row in range(first_row, last_row + 1)
        if ws[f"{unit_column}{row}"].value not in (None, "")
    }

    totals: dict[tuple[str, str], float] = defaultdict(float)
    for record in mapped_records:
        totals[(record.unit, record.field)] += record.area

    written = []
    missing_targets = []
    for (unit, field), area in totals.items():
        target_row = row_by_unit.get(unit)
        target_field = fields.get(field)
        if target_row is None or target_field is None:
            missing_targets.append({"unit": unit, "field": field, "area": area})
            continue
        cell = f"{target_field['column']}{target_row}"
        ws[cell] = round(area, 4)
        written.append({"cell": cell, "unit": unit, "field": field, "area": area})

    wb.save(output_path)
    return {"written": written, "missing_targets": missing_targets}
