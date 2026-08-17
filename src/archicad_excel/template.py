from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def load_yaml(path: str | Path) -> dict[str, Any]:
    text = Path(path).read_text(encoding="utf-8")
    try:
        import yaml

        data = yaml.safe_load(text) or {}
    except ModuleNotFoundError:
        data = _load_simple_yaml(text)
    if not isinstance(data, dict):
        raise ValueError(f"Expected mapping at root of {path}")
    return data


def _load_simple_yaml(text: str) -> dict[str, Any]:
    """Small YAML subset loader for public example configs.

    Supports nested mappings and scalar lists. It is intentionally limited;
    install PyYAML for full YAML support.
    """

    lines = text.splitlines()
    root = {}
    stack = [(-1, root)]
    for index, raw_line in enumerate(lines):
        if not raw_line.strip() or raw_line.lstrip().startswith("#"):
            continue
        indent = len(raw_line) - len(raw_line.lstrip(" "))
        line = raw_line.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()
        parent = stack[-1][1]
        if line.startswith("- "):
            if not isinstance(parent, list):
                raise ValueError("Invalid simple YAML list placement")
            parent.append(_parse_scalar(line[2:].strip()))
            continue
        key, _, value_text = line.partition(":")
        key = key.strip()
        value_text = value_text.strip()
        if value_text:
            parent[key] = _parse_scalar(value_text)
            continue
        child_is_list = _next_content_startswith(lines, index, "- ")
        child = [] if child_is_list else {}
        parent[key] = child
        stack.append((indent, child))
    return root


def _parse_scalar(value: str) -> Any:
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        value = value[1:-1]
    return value.encode("utf-8").decode("unicode_escape") if "\\t" in value else value


def _next_content_startswith(lines: list[str], current_index: int, prefix: str) -> bool:
    current_line = lines[current_index]
    current_indent = len(current_line) - len(current_line.lstrip(" "))
    for line in lines[current_index + 1 :]:
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        indent = len(line) - len(line.lstrip(" "))
        return indent > current_indent and line.strip().startswith(prefix)
    return False


def inspect_template(path: str | Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=False)
    result: dict[str, Any] = {"worksheets": []}
    for ws in wb.worksheets:
        formulas = []
        rotated_headers = []
        non_empty = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty.append(cell.coordinate)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
                if cell.alignment and cell.alignment.textRotation:
                    rotated_headers.append({"cell": cell.coordinate, "text": cell.value})

        result["worksheets"].append(
            {
                "name": ws.title,
                "used_range": f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_cells": [str(rng) for rng in ws.merged_cells.ranges],
                "print_area": ws.print_area,
                "rotated_headers": rotated_headers[:50],
                "formulas": formulas[:100],
                "candidate_headers": _candidate_headers(ws),
                "candidate_total_rows": _candidate_total_rows(ws),
                "non_empty_sample": non_empty[:50],
            }
        )
    return result


def _candidate_headers(ws) -> list[dict[str, Any]]:
    candidates = []
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        text_count = sum(1 for value in values if isinstance(value, str) and value.strip())
        if text_count >= 3:
            candidates.append({"row": row_idx, "values": [str(v) if v is not None else "" for v in values]})
    return candidates[:10]


def _candidate_total_rows(ws) -> list[dict[str, Any]]:
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        formulas = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col).value
            if isinstance(value, str) and value.upper().startswith("=SUM("):
                formulas.append(ws.cell(row_idx, col).coordinate)
        if formulas:
            rows.append({"row": row_idx, "sum_formula_cells": formulas})
    return rows[:20]
